import openpyxl
import os
import json
rowData = []
EXPORT_WITH_LABELS = False
column_order = [
    "Surprise %",
    "Mean",
    "SmartEstimate®",
    "Predicted Surprise",
    "Predicted Surprise %",
    "Median",
    "High",
    "Low",
    "Standard Deviation",
    "Standardized Unexpected Earnings (SUE)",
    "Number of Estimates",
    "YoY Growth %",
    "YoY Growth",
    "Mean % Chg (7d Post Rpt)"
]
for filename in os.listdir('data'):
    fileData = []
    wrkbk = openpyxl.load_workbook('data/' + filename)
    companyName = filename.split(' ')[2].split('.')[0]
    sh = wrkbk.active
    # First step. Get all column names.
    columns = {}
    itemNames = []
    for i in range(11, sh.max_row + 1):
        columns[sh.cell(row=i, column=1).value] = []
        itemNames.append(sh.cell(row=i, column=1).value)
    for j in range(3, sh.max_column-5):
        for i in range(11, sh.max_row + 1): 
            item = sh.cell(row=i, column=j).value
            if(item == '-'):
                item = 0
            columns[itemNames[i - 11]].append(item)
    rowData += [[companyName, columns]]
if(EXPORT_WITH_LABELS):
    # Use this for data analysis
    with open("database.json", "w") as outfile:
        json.dump(rowData, outfile, indent=4)
    with open("miniDatabase.json", "w") as outfile:
        json.dump(rowData[0:100], outfile, indent=4)
else:
    # Raw data, more useful for regressiob
    totalData = []
    for datum in rowData:
        usefulDatum = datum[1]
        for i in range(0, len(usefulDatum[column_order[0]])):
            subData = []
            for order in column_order:
                subData.append(usefulDatum[order][i])
            totalData.append([subData, usefulDatum["7d Price Reaction"][i]])
    with open("database.json", "w") as outfile:
        json.dump(totalData, outfile, indent=4)
    with open("miniDatabase.json", "w") as outfile:
        json.dump(totalData[0:100], outfile, indent=4)